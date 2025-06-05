import sys
import re
from openai import OpenAI
from openpyxl import load_workbook

EXAMPLE_FILE = './dataset/example_10'
API_KEY = 'your api key for OpenAI'
MODEL_KEY = 'your fine-tuned model key'

def run_sysgpt(client: OpenAI, problem_observation: str, model_key: str, temp: float):
    completion = client.chat.completions.create(
        model= model_key,
        temperature = temp,
        messages=[
            {
                "role": "system",
                "content": "You are an expert in Computer Science, especially in Systems area, who explains things specifically and comprehensively. You know the following categories that are common methodologies to improve system performance in a single-line execution, excluding those benefited from parallelism and algorithmic optimizations: 1.  Batching: Merge duplicate costs by grouping data or operations; 2. Caching: Memorize computed result and reuse it to avoid redundant computation; 3. Precomputing: Conduct initialization or execution in advance; 4. Deferring: Delay initialization or execution until it is needed or it has better context to make decision; 5. Relaxation: Cut workload size by sacrificing accuracy with approximation; 6. Contextualization: Collect additional data at runtime to make better decisions; 7. Hardware: Utilize specific hardware features, e.g., NUMA, NVM, FPGA, SmartNIC, to optimize workload computation; 8. Bypass: Skip existing layer by taking a fast path; 9. Delayering: Merge multiple layers into one to avoid intermediate costs among layers; 10. Decoupling: Split one layer into multiple layers to have finer control. Given problem description with observations, provide a system solution for improving performance to the problem. Explain the solution in detail using the methodologies described above."
            },
            {
                "role": "user",
                "content": "Given problem description with observations, provide **solutions** for improving performance to the problem. Give solutions in bullet points. Each bullet item must have following format:\n\n`1. [Methodology] Expanded description of this optimization technique.`\n\n The **Methodology** must be selected from one of the 10 methodologies."
                    f"Here is the problem description with observations:\n```\n{problem_observation}\n```\n"
            }
        ]
    )

    # print(completion.choices[0].message.content)
    return completion.choices[0].message.content

def run_few_shot(client: OpenAI, problem_observation: str, examples: str, temp: float):
    completion = client.chat.completions.create(
        model="gpt-4o-2024-08-06",
        temperature = temp,
        messages=[
            {
                "role": "system",
                "content": "You are an expert in Computer Science, especially in Systems area, who explains things specifically and comprehensively. You know the following categories that are common methodologies to improve system performance in a single-line execution, excluding those benefited from parallelism and algorithmic optimizations: 1.  Batching: Merge duplicate costs by grouping data or operations; 2. Caching: Memorize computed result and reuse it to avoid redundant computation; 3. Precomputing: Conduct initialization or execution in advance; 4. Deferring: Delay initialization or execution until it is needed or it has better context to make decision; 5. Relaxation: Cut workload size by sacrificing accuracy with approximation; 6. Contextualization: Collect additional data at runtime to make better decisions; 7. Hardware: Utilize specific hardware features, e.g., NUMA, NVM, FPGA, SmartNIC, to optimize workload computation; 8. Bypass: Skip existing layer by taking a fast path; 9. Delayering: Merge multiple layers into one to avoid intermediate costs among layers; 10. Decoupling: Split one layer into multiple layers to have finer control. Given problem description with observations, provide a system solution for improving performance to the problem. Explain the solution in detail using the methodologies described above."
            },
            {
                "role": "user",
                "content": "Given problem description with observations, provide **solutions** for improving performance to the problem. Give solutions in bullet points. Each bullet item must have following format:\n\n`1. [Methodology] Expanded description of this optimization technique.`\n\n The **Methodology** must be selected from one of the 10 methodologies."
                    f"{examples}"
                    f"Here is the problem description with observations:\n```\n{problem_observation}\n```\n"
            }
        ]
    )
    # print(completion.choices[0].message.content)
    return completion.choices[0].message.content


def extract_and_convert_methodologies(text):
    allowed = [
        "Batching", "Caching", "Precomputing", "Deferring", "Relaxation",
        "Contextualization", "Hardware", "Bypass", "Delayering", "Decoupling"
    ]

    normalization_map = {
        "bypassing": "Bypass",
        "bypass": "Bypass",
        "deferring": "Deferring",
        "decoupling": "Decoupling",
        "delayering": "Delayering",
        "caching": "Caching",
        "batching": "Batching",
        "precomputing": "Precomputing",
        "contextualization": "Contextualization",
        "relaxation": "Relaxation",
        "hardware": "Hardware"
    }

    # Extract [Methodology]
    tags = re.findall(r'\[(.*?)\]', text)

    # Check validity
    normalized_tags = []
    for tag in tags:
        key = tag.lower()
        if key in normalization_map:
            normalized_tags.append(normalization_map[key])
        else:
            return False

    final_vector = [0] * len(allowed)


    for tag in normalized_tags:
        index = allowed.index(tag)
        final_vector[index] = 1

    return final_vector


def calculate_metrics(ground_truth, prediction):
    assert len(ground_truth) == len(prediction), "Lists must be the same length"

    true_positive = sum((g == 1 and p == 1) for g, p in zip(ground_truth, prediction))
    false_positive = sum((g == 0 and p == 1) for g, p in zip(ground_truth, prediction))
    false_negative = sum((g == 1 and p == 0) for g, p in zip(ground_truth, prediction))

    precision = true_positive / (true_positive + false_positive) if (true_positive + false_positive) else 0
    recall = true_positive / (true_positive + false_negative) if (true_positive + false_negative) else 0
    f1_score = (2 * precision * recall) / (precision + recall) if (precision + recall) else 0

    return precision, recall, f1_score


def run_test(client: OpenAI, temp: float, trial: int):
    example_file = open(EXAMPLE_FILE, 'r').read()

    # Open Excel dataset
    wb = load_workbook(filename='./dataset/dataset.xlsx')
    ws = wb['testset']

    methodology = [
        "Batching", "Caching", "Precomputing", "Deferring", "Relaxation",
        "Contextualization", "Hardware", "Bypass", "Delayering", "Decoupling"
    ]

    n = 0

    sysgpt_total_precision = 0
    sysgpt_total_recall = 0
    sysgpt_total_f1 = 0

    few_shot_total_precision = 0
    few_shot_total_recall = 0
    few_shot_total_f1 = 0
    few_shot_method_num = 0

    for row in range(2, 98):
        is_scope = ws['G'+str(row)].value

        if is_scope == 'x':
            continue
        elif is_scope == 'o':
            print(f'=== {row} ===')

            problem = ws['D'+str(row)].value
            observation = ws['E'+str(row)].value

            problem_observation = problem + '\n' + observation

            # Get label vector
            ground_truth = extract_and_convert_methodologies(ws['F'+str(row)].value)
            print(ground_truth)

            # SysGPT ===============

            best_f1 = -1
            best_precision = 0
            best_recall = 0

            for i in range(0, trial):
                # Get sysgpt predict vector
                while True:
                    sysgpt_answer = run_sysgpt(client, problem_observation, MODEL_KEY, temp)
                    sysgpt_prediction = extract_and_convert_methodologies(sysgpt_answer)
                    if isinstance(sysgpt_prediction, bool):
                        print(sysgpt_answer)
                        print("Extract Error!! Retry...")
                    else:
                        break

                print("> Try ", i+1)
                print(sysgpt_prediction)

                precision, recall, f1 = calculate_metrics(ground_truth, sysgpt_prediction)
                if f1 > best_f1:
                    best_f1 = f1
                    best_precision = precision
                    best_recall = recall


            ## update sysgpt precision
            sysgpt_total_precision += best_precision
            sysgpt_total_recall += best_recall
            sysgpt_total_f1 += best_f1
            n += 1

            # Baseline ===============

            best_f1 = -1
            best_precision = 0
            best_recall = 0
            best_num = 0

            # Get 3-shot predict vector
            for i in range(0, trial):
                while True:
                    few_shot = run_few_shot(client, problem_observation, example_file, temp)
                    few_shot_prediction = extract_and_convert_methodologies(few_shot)
                    if isinstance(few_shot_prediction, bool):
                        print(few_shot)
                        print("Extract Error!! Retry...")
                    else:
                        break

                print("> Try ", i + 1)
                print(few_shot_prediction)

                precision, recall, f1 = calculate_metrics(ground_truth, few_shot_prediction)
                if f1 > best_f1:
                    best_f1 = f1
                    best_precision = precision
                    best_recall = recall
                    best_num = sum(few_shot_prediction)


            # update 3-shot precision
            few_shot_total_precision += best_precision
            few_shot_total_recall += best_recall
            few_shot_total_f1 += best_f1
            few_shot_method_num += best_num

            print("sysgpt  f1: ", sysgpt_total_f1 / n)
            print("fewshot f1: ", few_shot_total_f1 / n)
            print("fewshot # : ", few_shot_method_num / n)


    print("=== sysgpt ===")
    print("avg_precision", sysgpt_total_precision / n)
    print("avg_recall", sysgpt_total_recall / n)
    print("avg_f1", sysgpt_total_f1 / n)

    print("=== few-shot ===")
    print("avg_precision", few_shot_total_precision / n)
    print("avg_recall", few_shot_total_recall / n)
    print("avg_f1", few_shot_total_f1 / n)



if __name__ == '__main__':
    client = OpenAI(api_key=API_KEY)

    TEMP = float(sys.argv[1])
    TRIAL = int(sys.argv[2])

    print("Test info:")
    print(" - TEMP     : ", TEMP)
    print(" - Baseline : ", EXAMPLE_FILE)
    print(" - # trials : ", TRIAL)
    print("\n\n")


    run_test(client, TEMP, TRIAL)
